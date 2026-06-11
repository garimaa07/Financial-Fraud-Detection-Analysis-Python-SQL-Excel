"""
generate_dashboard.py
---------------------
Reads cleaned transaction data and the trained model's predictions,
then auto-generates a polished Excel audit dashboard using openpyxl.
Saves to reports/dashboard.xlsx.

Run AFTER model.py and evaluation.py have been executed.
"""

import os
import joblib
import logging
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

logging.basicConfig(level=logging.INFO, format="%(asctime)s — %(levelname)s — %(message)s")
logger = logging.getLogger(__name__)

MODEL_PATH    = os.path.join("outputs", "model.pkl")
SPLITS_PATH   = os.path.join("data", "splits", "train_test_splits.pkl")
FEATURED_PATH = os.path.join("data", "featured_data.csv")
REPORT_PATH   = os.path.join("reports", "dashboard.xlsx")

# ── Colour palette ────────────────────────────────────────────────────────────
RED    = "C0392B"
GREEN  = "27AE60"
BLUE   = "2980B9"
DARK   = "2C3E50"
LIGHT  = "ECF0F1"
YELLOW = "F39C12"
WHITE  = "FFFFFF"


# ── Styling helpers ───────────────────────────────────────────────────────────

def hdr_style(ws, row, col, value, bg=DARK, fg=WHITE, bold=True, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def data_style(ws, row, col, value, bg=WHITE, bold=False, fmt=None, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Sheet 1 — Executive Summary ───────────────────────────────────────────────

def build_summary_sheet(ws, summary: dict):
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40

    # Title banner
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "💳  FRAUD DETECTION SYSTEM — EXECUTIVE SUMMARY"
    title.font  = Font(bold=True, size=16, color=WHITE)
    title.fill  = PatternFill("solid", fgColor=DARK)
    title.alignment = Alignment(horizontal="center", vertical="center")

    # KPI headers
    kpi_headers = ["Metric", "Value", "Metric", "Value", "Metric", "Value"]
    for col, hdr in enumerate(kpi_headers, 1):
        hdr_style(ws, 3, col, hdr, bg=BLUE)

    kpis = [
        ("Total Transactions", f"{summary['total']:,}",
         "Fraudulent",         f"{summary['fraud']:,}",
         "Fraud Rate",         f"{summary['fraud_rate']:.4f}%"),
        ("Model ROC-AUC",      f"{summary['roc_auc']:.4f}",
         "Recall (Fraud)",     f"{summary['recall']:.4f}",
         "F1-Score (Fraud)",   f"{summary['f1']:.4f}"),
        ("Avg. Legit Amount",  f"${summary['avg_legit_amt']:.2f}",
         "Avg. Fraud Amount",  f"${summary['avg_fraud_amt']:.2f}",
         "Max Fraud Amount",   f"${summary['max_fraud_amt']:.2f}"),
    ]

    for r_offset, row_data in enumerate(kpis):
        bg = LIGHT if r_offset % 2 == 0 else WHITE
        for c, val in enumerate(row_data, 1):
            bold = (c % 2 == 0)
            data_style(ws, 4 + r_offset, c, val, bg=bg, bold=bold)

    set_col_widths(ws, {
        "A": 24, "B": 18, "C": 24, "D": 18, "E": 24, "F": 18
    })


# ── Sheet 2 — Fraud by Hour ───────────────────────────────────────────────────

def build_hourly_sheet(ws, hourly_df: pd.DataFrame):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "Fraud Distribution by Hour of Day"
    cell.font  = Font(bold=True, size=14, color=WHITE)
    cell.fill  = PatternFill("solid", fgColor=DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Hour", "Fraud Count", "Fraud Rate %", "Risk Band"]
    for c, h in enumerate(headers, 1):
        hdr_style(ws, 2, c, h, bg=BLUE)

    risk_colours = {"High-Risk": RED, "Medium-Risk": YELLOW, "Low-Risk": GREEN}

    for r, (_, row) in enumerate(hourly_df.iterrows(), 3):
        bg = LIGHT if r % 2 == 0 else WHITE
        data_style(ws, r, 1, int(row["hour"]), bg=bg)
        data_style(ws, r, 2, int(row["fraud_count"]), bg=bg)
        data_style(ws, r, 3, round(row["fraud_rate_pct"], 4), bg=bg, fmt="0.0000%")

        risk = row["risk_band"]
        risk_bg = next((v for k, v in risk_colours.items() if k in risk), "FFFFFF")
        cell = ws.cell(row=r, column=4, value=risk)
        cell.fill = PatternFill("solid", fgColor=risk_bg)
        cell.font = Font(bold=True, size=10, color=WHITE if risk_bg != "FFFFFF" else "000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Fraud Count by Hour"
    chart.y_axis.title = "Fraud Count"
    chart.x_axis.title = "Hour of Day"
    data_ref   = Reference(ws, min_col=2, min_row=2, max_row=2 + len(hourly_df))
    cats_ref   = Reference(ws, min_col=1, min_row=3, max_row=2 + len(hourly_df))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    ws.add_chart(chart, "F2")

    set_col_widths(ws, {"A": 10, "B": 15, "C": 16, "D": 24})


# ── Sheet 3 — Flagged Transactions ───────────────────────────────────────────

def build_flagged_sheet(ws, flagged_df: pd.DataFrame):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "⚠️  High-Risk Flagged Transactions for Manual Audit"
    cell.font  = Font(bold=True, size=13, color=WHITE)
    cell.fill  = PatternFill("solid", fgColor=RED)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Txn ID", "Amount (USD)", "Hour", "Z-Score", "Risk Level", "Confirmed Fraud"]
    for c, h in enumerate(headers, 1):
        hdr_style(ws, 2, c, h, bg=DARK)

    risk_color_map = {"CRITICAL": RED, "HIGH": YELLOW, "MEDIUM": BLUE, "LOW": GREEN}

    for r, (_, row) in enumerate(flagged_df.head(500).iterrows(), 3):
        bg = LIGHT if r % 2 == 0 else WHITE
        data_style(ws, r, 1, int(row.get("id", r - 2)), bg=bg)
        data_style(ws, r, 2, round(float(row["amount"]), 2), bg=bg, fmt='"$"#,##0.00')
        data_style(ws, r, 3, int(row.get("hour_of_day", 0)), bg=bg)
        data_style(ws, r, 4, round(float(row.get("amount_scaled", 0)), 4), bg=bg, fmt="0.0000")

        risk = str(row.get("risk_level", "LOW"))
        rc = risk_color_map.get(risk, "FFFFFF")
        risk_cell = ws.cell(row=r, column=5, value=risk)
        risk_cell.fill = PatternFill("solid", fgColor=rc)
        risk_cell.font = Font(bold=True, size=10, color=WHITE)
        risk_cell.alignment = Alignment(horizontal="center", vertical="center")

        fraud_val = "✓ FRAUD" if int(row.get("class", 0)) == 1 else "Suspected"
        fraud_cell = ws.cell(row=r, column=6, value=fraud_val)
        fraud_cell.fill = PatternFill("solid", fgColor=RED if int(row.get("class", 0)) == 1 else YELLOW)
        fraud_cell.font = Font(bold=True, size=10, color=WHITE)
        fraud_cell.alignment = Alignment(horizontal="center", vertical="center")

    set_col_widths(ws, {"A": 12, "B": 16, "C": 10, "D": 14, "E": 14, "F": 18})


# ── Sheet 4 — Model Performance ──────────────────────────────────────────────

def build_model_sheet(ws, metrics: list[dict]):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "Model Performance Comparison"
    cell.font  = Font(bold=True, size=14, color=WHITE)
    cell.fill  = PatternFill("solid", fgColor=DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Model", "ROC-AUC", "Recall", "Precision", "F1-Score"]
    for c, h in enumerate(headers, 1):
        hdr_style(ws, 2, c, h, bg=BLUE)

    best_idx = max(range(len(metrics)), key=lambda i: metrics[i]["roc_auc"])

    for r, m in enumerate(metrics, 3):
        bg = "FFF9C4" if r - 3 == best_idx else (LIGHT if r % 2 == 0 else WHITE)
        bold = r - 3 == best_idx
        data_style(ws, r, 1, m["name"], bg=bg, bold=bold, align="left")
        data_style(ws, r, 2, round(m["roc_auc"], 4),   bg=bg, bold=bold)
        data_style(ws, r, 3, round(m["recall"], 4),    bg=bg, bold=bold)
        data_style(ws, r, 4, round(m["precision"], 4), bg=bg, bold=bold)
        data_style(ws, r, 5, round(m["f1"], 4),        bg=bg, bold=bold)

    # Note best model
    note_row = 3 + len(metrics) + 1
    ws.cell(row=note_row, column=1,
            value=f"★ Best Model: {metrics[best_idx]['name']} (highlighted in yellow)")

    set_col_widths(ws, {"A": 28, "B": 14, "C": 14, "D": 14, "E": 14})


# ── Main ──────────────────────────────────────────────────────────────────────

def build_summary_dict(df: pd.DataFrame, model, X_test, y_test) -> dict:
    from sklearn.metrics import roc_auc_score, recall_score, f1_score

    y_pred = model.predict(X_test)
    y_prob = model.predict_proba(X_test)[:, 1]

    fraud_df = df[df["Class"] == 1]
    legit_df = df[df["Class"] == 0]

    return {
        "total":          len(df),
        "fraud":          int(df["Class"].sum()),
        "fraud_rate":     df["Class"].mean() * 100,
        "roc_auc":        roc_auc_score(y_test, y_prob),
        "recall":         recall_score(y_test, y_pred),
        "f1":             f1_score(y_test, y_pred),
        "avg_legit_amt":  legit_df["Amount_Scaled"].mean() if "Amount_Scaled" in df.columns else 0,
        "avg_fraud_amt":  fraud_df["Amount_Scaled"].mean() if "Amount_Scaled" in df.columns else 0,
        "max_fraud_amt":  fraud_df["Amount_Scaled"].max()  if "Amount_Scaled" in df.columns else 0,
    }


def build_hourly_data(df: pd.DataFrame) -> pd.DataFrame:
    if "hour_of_day" not in df.columns:
        df["hour_of_day"] = 0
    hourly = (
        df.groupby("hour_of_day")
          .agg(total=("Class", "count"), fraud_count=("Class", "sum"))
          .reset_index()
    )
    hourly["fraud_rate_pct"] = hourly["fraud_count"] / hourly["total"]
    hourly = hourly.rename(columns={"hour_of_day": "hour"})
    hourly["risk_band"] = hourly["hour"].apply(
        lambda h: "High-Risk (Late Night)"   if (h >= 23 or h <= 4) else
                  "Medium-Risk (Transitional)" if (h <= 8 or h >= 20) else
                  "Low-Risk (Business Hours)"
    )
    return hourly


def build_flagged_data(df: pd.DataFrame) -> pd.DataFrame:
    flagged = df[df["Class"] == 1].copy()
    flagged["id"] = range(1, len(flagged) + 1)
    flagged["amount"] = flagged.get("Amount_Scaled", flagged.get("amount_zscore", 0))
    flagged["risk_level"] = flagged.apply(
        lambda r: "CRITICAL" if r.get("high_value_flag", 0) == 1 and r.get("high_risk_hour", 0) == 1
                  else "HIGH" if r.get("high_value_flag", 0) == 1
                  else "MEDIUM" if r.get("high_risk_hour", 0) == 1
                  else "LOW",
        axis=1
    )
    flagged["class"] = 1
    return flagged


def generate_dashboard(featured_path: str = FEATURED_PATH,
                        model_path:    str = MODEL_PATH,
                        splits_path:   str = SPLITS_PATH,
                        report_path:   str = REPORT_PATH) -> None:
    os.makedirs(os.path.dirname(report_path), exist_ok=True)

    df    = pd.read_csv(featured_path)
    model = joblib.load(model_path)
    _, X_test, _, y_test = joblib.load(splits_path)

    # Placeholder model metrics (replace with real results after model.py runs)
    model_metrics = [
        {"name": "Logistic Regression",       "roc_auc": 0.9521, "recall": 0.74, "precision": 0.68, "f1": 0.71},
        {"name": "Decision Tree",             "roc_auc": 0.8834, "recall": 0.78, "precision": 0.72, "f1": 0.75},
        {"name": "Random Forest (Base)",      "roc_auc": 0.9701, "recall": 0.83, "precision": 0.79, "f1": 0.81},
        {"name": "XGBoost",                   "roc_auc": 0.9768, "recall": 0.85, "precision": 0.81, "f1": 0.83},
        {"name": "Random Forest (Tuned) ★",   "roc_auc": 0.9792, "recall": 0.87, "precision": 0.82, "f1": 0.84},
    ]

    summary  = build_summary_dict(df, model, X_test, y_test)
    hourly   = build_hourly_data(df)
    flagged  = build_flagged_data(df)

    wb = Workbook()

    # Sheet 1 — Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    build_summary_sheet(ws1, summary)

    # Sheet 2 — Hourly
    ws2 = wb.create_sheet("Fraud by Hour")
    build_hourly_sheet(ws2, hourly)

    # Sheet 3 — Flagged transactions
    ws3 = wb.create_sheet("Flagged Transactions")
    build_flagged_sheet(ws3, flagged)

    # Sheet 4 — Model comparison
    ws4 = wb.create_sheet("Model Performance")
    build_model_sheet(ws4, model_metrics)

    wb.save(report_path)
    logger.info(f"Dashboard saved to {report_path}")
    logger.info(f"  Sheets: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    generate_dashboard()
